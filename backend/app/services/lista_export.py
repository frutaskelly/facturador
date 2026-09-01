"""Una lista de precios que sale y regresa: Excel de ida y vuelta, y PDF.

Pedido del dueño (28-ago-2026): desde la pantalla de Precios poder (a) bajar
el PDF de la lista para mandarla, (b) bajar un Excel con los precios y
(c) subir ESE MISMO Excel para actualizar en masa — agregar renglones nuevos,
cambiar precios o quitar renglones.

El contrato del Excel (mismas columnas al exportar y al importar):

    SKU | PRODUCTO | PRESENTACION | DESDE CANTIDAD | PRECIO

- El renglón se identifica por (SKU, PRESENTACION, DESDE CANTIDAD).
- PRECIO con valor → se crea o se actualiza.
- PRECIO vacío o 0 → el renglón SE QUITA de la lista (así se poda sin borrar
  a mano). PRODUCTO es informativo: el que manda es el SKU.
- Un SKU que no existe en el catálogo se reporta como error de esa fila y las
  demás siguen — aquí no se dan de alta productos (para eso está el wizard de
  importación, que crea producto + código del cliente + todo lo demás).
"""
from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from uuid import UUID

from sqlalchemy.orm import Session

from ..models import ListaPrecios, Precio, Producto
from .inventario import presentacion_declarada

HDR = ["SKU", "PRODUCTO", "PRESENTACION", "DESDE CANTIDAD", "PRECIO"]


def _filas(db: Session, lista: ListaPrecios) -> list[tuple]:
    q = (
        db.query(Precio, Producto.sku, Producto.nombre)
        .join(Producto, Producto.id == Precio.producto_id)
        .filter(Precio.lista_id == lista.id, Producto.deleted_at.is_(None))
        .order_by(Producto.nombre.asc(), Precio.presentacion.asc(), Precio.cantidad_minima.asc())
    )
    return [(sku, nombre, p.presentacion, p.cantidad_minima, p.precio_unitario)
            for p, sku, nombre in q.all()]


def exportar_xlsx(db: Session, lista: ListaPrecios) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Precios"
    ws.append(HDR)
    for sku, nombre, pres, cant, precio in _filas(db, lista):
        ws.append([sku, nombre, pres, float(cant), float(precio)])
    # anchos legibles: nadie quiere reacomodar columnas antes de trabajar
    for col, ancho in zip("ABCDE", (14, 46, 14, 16, 12)):
        ws.column_dimensions[col].width = ancho
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_pdf(db: Session, lista: ListaPrecios, tenant) -> bytes:
    """La lista con el MEMBRETE del negocio (layout Smart Supply, 29-ago-2026):
    logo + datos fiscales, tabla azul con zebra y folio de página."""
    from .reporte_pdf import CELDA, construir, membrete, tabla_reporte
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph

    filas = []
    for _sku, nombre, pres, cant, precio in _filas(db, lista):
        etiqueta = pres if cant in (1, Decimal("1")) else f"{pres} (desde {cant})"
        filas.append([Paragraph(nombre, CELDA), etiqueta, f"${Decimal(precio):,.2f}"])
    partes = membrete(tenant, "Lista de precios",
                      f"{lista.nombre} · {len(filas)} productos")
    partes.append(tabla_reporte(
        ["Producto", "Presentación", "Precio"], filas,
        [108 * mm, 38 * mm, 32 * mm], num_cols=(2,),
    ))
    return construir(f"Lista de precios · {lista.nombre}", partes)



def importar_xlsx(db: Session, tenant_id: UUID, lista: ListaPrecios, data: bytes) -> dict:
    """Aplica el Excel de ida y vuelta. Devuelve el resumen + errores por fila."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas or [str(x or "").strip().upper() for x in filas[0][:5]] != HDR:
        return {"ok": False, "error": "El archivo no trae las columnas esperadas "
                                      f"({' | '.join(HDR)}) — baja el Excel de la lista y edítalo"}

    skus = {}
    prods = {}
    for p in db.query(Producto).filter(Producto.tenant_id == tenant_id, Producto.deleted_at.is_(None)):
        skus[(p.sku or "").strip()] = p.id
        prods[p.id] = p

    existentes = {
        (pr.producto_id, pr.presentacion, Decimal(pr.cantidad_minima)): pr
        for pr in db.query(Precio).filter(Precio.lista_id == lista.id)
    }

    res = {"ok": True, "actualizados": 0, "agregados": 0, "eliminados": 0,
           "sin_cambio": 0, "errores": []}
    import uuid as _uuid
    for i, fila in enumerate(filas[1:], start=2):
        sku = str(fila[0] or "").strip()
        if not sku:
            continue
        pres = str(fila[2] or "").strip().upper() or "KILO"
        try:
            cant = Decimal(str(fila[3] if fila[3] not in (None, "") else 1))
        except InvalidOperation:
            res["errores"].append(f"fila {i}: DESDE CANTIDAD ilegible"); continue
        crudo = fila[4]
        pid = skus.get(sku)
        if pid is None:
            res["errores"].append(f"fila {i}: el SKU {sku} no existe en el catálogo "
                                  "(los productos nuevos se dan de alta en Productos → Importar)")
            continue
        # Misma regla que el alta por pantalla: un precio en una presentación
        # que el producto no declara no lo cobra nadie, y aquí llegan archivos
        # editados a mano donde escribir CAJA en vez de KILO es un teclazo.
        if not presentacion_declarada(prods.get(pid), pres):
            res["errores"].append(
                f"fila {i}: {prods[pid].nombre} no maneja la presentación {pres} "
                "(agrégasela al producto, con cuántas unidades base trae)")
            continue
        llave = (pid, pres, cant)
        actual = existentes.get(llave)
        if crudo in (None, "") or (isinstance(crudo, (int, float, Decimal)) and Decimal(str(crudo)) == 0):
            if actual is not None:
                db.delete(actual)
                res["eliminados"] += 1
            continue
        try:
            precio = Decimal(str(crudo)).quantize(Decimal("0.0001"))
            if precio < 0:
                raise InvalidOperation
        except InvalidOperation:
            res["errores"].append(f"fila {i}: PRECIO ilegible ({crudo!r})"); continue
        if actual is None:
            db.add(Precio(id=_uuid.uuid4(), tenant_id=tenant_id, lista_id=lista.id,
                          producto_id=pid, presentacion=pres,
                          precio_unitario=precio, cantidad_minima=cant))
            existentes[llave] = None   # evita duplicar si el archivo repite la fila
            res["agregados"] += 1
        elif Decimal(actual.precio_unitario) != precio:
            actual.precio_unitario = precio
            res["actualizados"] += 1
        else:
            res["sin_cambio"] += 1
    db.flush()
    return res
