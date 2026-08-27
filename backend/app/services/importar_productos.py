"""Importación masiva de productos: plantilla propia o lista de precios ajena.

Dos caminos, un mismo resultado (filas estructuradas para el preview):

1. **Plantilla** (xlsx/csv con nuestros encabezados) → parseo determinista, sin
   IA. Es el formato que descarga el usuario desde /productos.
2. **Cualquier lista de precios** (Excel/CSV con otro acomodo, PDF o foto) → la
   IA extrae {nombre, codigo, descripcion, unidad, precio} y de paso sugiere
   clave/unidad SAT por fila (una sola llamada, no una por producto).

El cruce contra el catálogo (para NO duplicar productos) y el alta viven en el
router; aquí solo archivo → filas.
"""
from __future__ import annotations

import base64
import io
import logging
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import Optional

from ..core.config import settings

logger = logging.getLogger(__name__)

MAX_FILAS = 2000         # tope del camino determinista (plantilla/SAE)
MAX_FILAS_IA = 500       # tope del camino IA (el output de la extracción cuesta)

# Encabezados de la plantilla oficial (en este orden). Compatible con el
# export de listas de SAE (CLAVE | DESCRIPCIÓN | UNIDAD DE SALIDA | CLAVE SAT |
# UNIDAD DE SALIDA SAT | PRECIO | CATEGORÍA | ESTATUS).
PLANTILLA_COLUMNAS = [
    ("DESCRIPCION", "Nombre del producto (obligatorio)"),
    ("CLAVE", "Código del cliente o SKU deseado (opcional)"),
    ("UNIDAD", "Unidad de venta: KILO, PIEZA, CAJA… (opcional, default KILO)"),
    ("PRECIO", "Precio de venta (opcional — crea/actualiza lista de precios)"),
    ("CLAVE_SAT", "Clave SAT c_ClaveProdServ, 8 dígitos (opcional — sugerida o genérica 01010101)"),
    ("UNIDAD_SAT", "Clave SAT de unidad: KGM, H87, XBX… (opcional — sugerida o genérica)"),
    ("CATEGORIA", "Categoría (opcional — se puede crear al importar)"),
    ("ESQUEMA_IMPUESTO", "Código del esquema de impuesto: IVA16, IVA0… (opcional)"),
    ("CODIGO_BARRAS", "Código de barras EAN/GTIN (opcional)"),
    ("ESTATUS", "ALTA o BAJA (opcional — BAJA se omite por default)"),
]


class ImportProductosError(Exception):
    """Error de formato/lectura con mensaje para el usuario."""


def _norm_header(h) -> str:
    s = unicodedata.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z]", "", s.upper())


# Encabezado normalizado → campo (tolera variantes).
_HEADERS = {
    "NOMBRE": "nombre",
    "PRODUCTO": "nombre",
    "DESCRIPCION": "descripcion",
    "CODIGO": "codigo",
    "CLAVE": "codigo",
    "CLAVESAE": "codigo",       # exports de SAE ASPEL
    "SKU": "codigo",
    "CLAVESKU": "codigo",
    "UNIDAD": "unidad",
    "UNIDADDESALIDA": "unidad",  # export SAE
    "UNIDADDEVENTA": "unidad",
    "PRESENTACION": "unidad",
    "PRECIO": "precio",
    "CLAVESAT": "clave_sat",
    "UNIDADSAT": "unidad_sat",
    "UNIDADDESALIDASAT": "unidad_sat",   # export SAE
    "CODIGOBARRAS": "codigo_barras",
    "CATEGORIA": "categoria",
    "LINEA": "categoria",
    "ESQUEMA": "esquema",
    "ESQUEMAIMPUESTO": "esquema",
    "ESQUEMADEIMPUESTO": "esquema",
    "ESTATUS": "estatus",
    "ESTADO": "estatus",
    # Columnas informativas de los exports que NO deben capturarse:
    # DESCRIPCIONSAT y DESCRIPCIONUNIDADSAT no están aquí a propósito.
}


def _texto(v) -> str:
    s = str(v if v is not None else "").strip()
    return "" if s.lower() in ("", "nan", "none") else s


def _decimal(v) -> Optional[Decimal]:
    """Precio del archivo → Decimal, tolerando formato local.

    La coma puede ser separador de MILES ("1,234.56") o DECIMAL ("12,50",
    exports en formato europeo). Borrar la coma a ciegas convertía 12,50 en
    1250 — un precio x100 que pasaba sin marca. Se decide por la forma:
    si hay punto y coma, manda el que va más a la derecha; si solo hay comas,
    es decimal cuando deja 1-2 dígitos al final y no hay grupos de 3.
    """
    s = _texto(v).replace("$", "").replace(" ", "").replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        # El separador decimal es el ÚLTIMO que aparece; el otro es de miles.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        entero, _, resto = s.rpartition(",")
        # "1,234" / "1,234,567" → miles; "12,5" / "12,50" → decimal.
        s = s.replace(",", "") if (len(resto) == 3 and entero) else s.replace(",", ".")
    try:
        d = Decimal(s)
        return d if d >= 0 else None
    except InvalidOperation:
        return None


# Variantes de unidad de venta → unidad canónica del sistema.
_UNIDADES_CANON = {
    "KILOGRAMO": "KILO", "KILOGRAMOS": "KILO", "KILOS": "KILO", "KG": "KILO", "KGS": "KILO",
    "PZ": "PIEZA", "PZA": "PIEZA", "PZAS": "PIEZA", "PIEZAS": "PIEZA", "PZS": "PIEZA",
    "LT": "LITRO", "LTS": "LITRO", "L": "LITRO", "LITROS": "LITRO",
    "ML": "MILILITRO", "MILILITROS": "MILILITRO",
    "GR": "GRAMO", "GRS": "GRAMO", "GRAMOS": "GRAMO", "G": "GRAMO",
    "MJ": "MANOJO", "MANOJOS": "MANOJO",
    "CJ": "CAJA", "CJA": "CAJA", "CAJAS": "CAJA",
    "PAQ": "PAQUETE", "PAQUETES": "PAQUETE",
    "BOLSAS": "BOLSA", "COSTALES": "COSTAL", "BULTOS": "BULTO",
    "DOC": "DOCENA", "DOCENAS": "DOCENA", "MALLAS": "MALLA", "REJAS": "REJA",
}


def normalizar_unidad(u: str) -> str:
    s = (u or "").strip().upper()
    return _UNIDADES_CANON.get(s, s)


def _leer_tabla(data: bytes, filename: str):
    """xlsx/xls/csv → DataFrame de textos (header en la fila 0)."""
    import pandas as pd

    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    try:
        if ext == "csv":
            return pd.read_csv(io.BytesIO(data), header=0, dtype=str, sep=None, engine="python")
        return pd.read_excel(io.BytesIO(data), sheet_name=0, header=0, dtype=str)
    except Exception as exc:  # noqa: BLE001 — el error de la lib es críptico
        raise ImportProductosError(
            f"No se pudo leer el archivo '{filename}'. ¿Es un Excel (.xlsx/.xls) o CSV?"
        ) from exc


# Campos del sistema a los que se puede mapear una columna del archivo.
CAMPOS_MAPEABLES = [
    ("nombre", "Descripción / nombre del producto"),
    ("codigo", "Clave / SKU"),
    ("descripcion", "Descripción adicional"),
    ("unidad", "Unidad de venta"),
    ("precio", "Precio"),
    ("clave_sat", "Clave SAT (producto/servicio)"),
    ("unidad_sat", "Unidad SAT"),
    ("categoria", "Categoría"),
    ("esquema", "Esquema de impuesto"),
    ("codigo_barras", "Código de barras"),
    ("estatus", "Estatus (ALTA/BAJA)"),
]


def _detectar_columnas(df) -> dict[int, str]:
    """Encabezados del archivo → campo del sistema (mapeo automático)."""
    cols: dict[int, str] = {}
    for i, h in enumerate(df.columns):
        campo = _HEADERS.get(_norm_header(h))
        if campo and campo not in cols.values():
            cols[i] = campo
    if "nombre" not in cols.values():
        # Exports estilo SAE: no hay columna NOMBRE, pero la DESCRIPCION es el
        # nombre del producto (Linea | Clave SAE | Descripcion | Unidad | Precio).
        desc_idx = next((i for i, c in cols.items() if c == "descripcion"), None)
        if desc_idx is not None:
            cols[desc_idx] = "nombre"
    return cols


def analizar_columnas(data: bytes, filename: str) -> Optional[dict]:
    """Qué columna del archivo se leyó como qué campo, con valores de muestra.

    Es lo que el usuario revisa ANTES del preview: puede reasignar una columna
    a otro campo o dejar de importarla. Devuelve None si el archivo no es
    tabular reconocible (→ camino IA)."""
    df = _leer_tabla(data, filename)
    detectadas = _detectar_columnas(df)
    columnas = []
    for i, h in enumerate(df.columns):
        # Se escanean bastantes filas, no solo las primeras: una columna vacía
        # al inicio y con datos más abajo se veía como "sin datos" e invitaba a
        # descartarla.
        muestras: list[str] = []
        for v in df.iloc[:400, i].tolist():
            t = _texto(v)
            if t:
                muestras.append(t)
            if len(muestras) == 3:
                break
        columnas.append({
            "indice": i,
            "encabezado": _texto(h) or f"Columna {i + 1}",
            "campo": detectadas.get(i, ""),        # "" = no se importa
            "muestras": muestras,
        })
    return {
        "columnas": columnas,
        "campos": [{"valor": v, "etiqueta": e} for v, e in CAMPOS_MAPEABLES],
    }


def parsear_plantilla(
    data: bytes, filename: str, mapeo: Optional[dict[int, str]] = None
) -> Optional[list[dict]]:
    """Camino determinista: si el archivo trae al menos la columna NOMBRE (o
    PRODUCTO) en el encabezado, se parsea sin IA. Devuelve None si el archivo
    no se parece a la plantilla (→ probar con IA).

    `mapeo` (índice de columna → campo) lo manda el usuario desde la pantalla
    de columnas y GANA sobre la detección automática."""
    df = _leer_tabla(data, filename)

    # `mapeo is not None` (no `if mapeo`): un mapeo vacío es una decisión del
    # usuario — "no importar nada" — y NO debe caer en auto-detección.
    if mapeo is not None:
        # Solo columnas que existen en el archivo y campos conocidos; un campo
        # no puede venir de dos columnas (gana la primera).
        validos = {v for v, _ in CAMPOS_MAPEABLES}
        cols: dict[int, str] = {}
        for i, campo in sorted(mapeo.items()):
            if 0 <= i < len(df.columns) and campo in validos and campo not in cols.values():
                cols[i] = campo
    else:
        cols = _detectar_columnas(df)
    if "nombre" not in cols.values():
        if mapeo is not None:
            raise ImportProductosError(
                "Falta indicar qué columna trae la DESCRIPCIÓN (nombre) del producto"
            )
        return None

    filas: list[dict] = []
    saltadas = 0
    for _, row in df.iterrows():
        r = {campo: row.iloc[i] for i, campo in cols.items()}
        nombre = _texto(r.get("nombre"))
        if not nombre:
            # Filas vacías / totales — pero si la fila trae ALGO en otra
            # columna, es un renglón real que se está descartando: se cuenta
            # para avisarlo (antes desaparecían en silencio).
            if any(_texto(v) for v in row.tolist()):
                saltadas += 1
            continue
        precio = _decimal(r.get("precio"))
        categoria = _texto(r.get("categoria"))
        # "— sin categoría —" y similares cuentan como vacío.
        if "sin categoria" in unicodedata.normalize("NFKD", categoria.lower()).encode(
            "ascii", "ignore"
        ).decode():
            categoria = ""
        filas.append({
            "nombre": nombre,
            "codigo": _texto(r.get("codigo")),
            "descripcion": _texto(r.get("descripcion")),
            "unidad": normalizar_unidad(_texto(r.get("unidad"))),
            "precio": str(precio) if precio is not None else "",
            "clave_sat": re.sub(r"\D", "", _texto(r.get("clave_sat")))[:8],
            "unidad_sat": _texto(r.get("unidad_sat")).upper()[:3],
            "codigo_barras": _texto(r.get("codigo_barras")),
            "categoria": categoria,
            "esquema": _texto(r.get("esquema")),
            "estatus": _texto(r.get("estatus")).upper(),
        })
        if len(filas) > MAX_FILAS:
            raise ImportProductosError(f"Máximo {MAX_FILAS} productos por archivo")
    if not filas:
        raise ImportProductosError("El archivo no tiene filas con NOMBRE de producto")
    # El contador viaja en la primera fila (el llamador lo lee y lo quita).
    filas[0]["_saltadas"] = saltadas
    return filas


# ─── Extracción con IA (lista de precios en cualquier formato) ───────────────

_MIME_POR_EXT = {
    "pdf": "application/pdf",
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "webp": "image/webp",
    "gif": "image/gif",
}

_SYSTEM_EXTRACT = """\
Eres un asistente que convierte LISTAS DE PRECIOS de frutas, verduras y \
abarrotes (en cualquier acomodo: Excel pegado, PDF o foto) en filas \
estructuradas de productos para darlos de alta en un catálogo mexicano (CFDI 4.0).

Reglas:
- UNA fila de salida por CADA renglón de producto del documento, en orden. \
NUNCA fusiones ni saltes renglones parecidos: si el documento repite un nombre \
(o solo cambia variedad, unidad o precio), devuelve cada renglón por separado.
- Ignora encabezados, títulos, totales, notas y filas vacías.
- `nombre`: el nombre del producto tal como viene (limpio de espacios dobles).
- `codigo`: el código/clave del producto SI la lista lo trae (p. ej. \
"JIT-SAD-001", "A0125"); vacío si no hay.
- `descripcion`: presentación/marca/detalle extra si viene aparte del nombre.
- `unidad`: unidad de venta si se distingue (KILO, PIEZA, CAJA, BULTO, LITRO, \
MANOJO…); vacío si no.
- `precio`: número sin símbolo (1234.50); vacío si la lista no trae precio.
- `clave_sat`: tu mejor sugerencia de c_ClaveProdServ (8 dígitos exactos) para \
ese producto. Frutas frescas 503xxxxx, verduras frescas 504xxxxx, abarrotes \
50xxxxxx; si dudas usa la categoría más cercana.
- `unidad_sat`: KGM (kilogramo), H87 (pieza), XBX (caja), LTR (litro), \
GRM (gramo), XPK (paquete), XBG (bolsa) — según la unidad de venta; KGM si dudas.
- NO inventes productos ni precios; si un dato no está, déjalo vacío.
Llama SIEMPRE a la herramienta `registrar_productos` con TODAS las filas.\
"""

_TOOL_EXTRACT = {
    "name": "registrar_productos",
    "description": "Registra las filas de productos extraídas de la lista de precios.",
    "input_schema": {
        "type": "object",
        "properties": {
            "productos": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "nombre": {"type": "string"},
                        "codigo": {"type": "string"},
                        "descripcion": {"type": "string"},
                        "unidad": {"type": "string"},
                        "precio": {"type": "string"},
                        "clave_sat": {"type": "string"},
                        "unidad_sat": {"type": "string"},
                        "categoria": {"type": "string", "description": "Categoría/línea si la lista la trae; vacío si no."},
                    },
                    "required": ["nombre"],
                },
            },
        },
        "required": ["productos"],
    },
}


def _tabla_a_texto(data: bytes, filename: str) -> str:
    """Excel/CSV con acomodo libre → texto tabular (TSV) para la IA."""
    import pandas as pd

    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    try:
        if ext == "csv":
            df = pd.read_csv(io.BytesIO(data), header=None, dtype=str, sep=None, engine="python")
        else:
            df = pd.read_excel(io.BytesIO(data), sheet_name=0, header=None, dtype=str)
    except Exception as exc:  # noqa: BLE001
        raise ImportProductosError(
            f"No se pudo leer el archivo '{filename}'. ¿Es un Excel (.xlsx/.xls) o CSV?"
        ) from exc
    lineas = []
    for _, row in df.iterrows():
        celdas = [_texto(v) for v in row.tolist()]
        if any(celdas):
            lineas.append("\t".join(celdas))
        if len(lineas) > MAX_FILAS_IA + 50:   # margen para encabezados/notas
            raise ImportProductosError(f"Máximo {MAX_FILAS_IA} productos por archivo con IA")
    if not lineas:
        raise ImportProductosError("El archivo está vacío")
    return "\n".join(lineas)


def extraer_con_ia(data: bytes, filename: str) -> list[dict]:
    """Lista de precios en cualquier formato → filas estructuradas vía Claude.

    Excel/CSV van como texto tabular; PDF como documento; imágenes como imagen.
    Lanza ImportProductosError con mensaje claro si la IA no está disponible.
    """
    if not settings.ANTHROPIC_API_KEY:
        raise ImportProductosError(
            "El archivo no coincide con la plantilla y la IA no está configurada. "
            "Descarga la plantilla y captura ahí los productos."
        )
    try:
        import anthropic
    except ImportError as exc:  # pragma: no cover
        raise ImportProductosError("SDK de IA no instalado") from exc

    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    if ext in ("xlsx", "xls", "csv"):
        content: list[dict] = [{"type": "text", "text":
            "Lista de precios (texto tabular, columnas separadas por tabulador):\n\n"
            + _tabla_a_texto(data, filename)}]
    elif ext == "pdf":
        content = [
            {"type": "document", "source": {
                "type": "base64", "media_type": "application/pdf",
                "data": base64.standard_b64encode(data).decode(),
            }},
            {"type": "text", "text": "Extrae los productos de esta lista de precios."},
        ]
    elif ext in _MIME_POR_EXT:  # imágenes
        content = [
            {"type": "image", "source": {
                "type": "base64", "media_type": _MIME_POR_EXT[ext],
                "data": base64.standard_b64encode(data).decode(),
            }},
            {"type": "text", "text": "Extrae los productos de esta lista de precios."},
        ]
    else:
        raise ImportProductosError(
            "Formato no soportado. Sube un Excel (.xlsx/.xls), CSV, PDF o foto (JPG/PNG)."
        )

    client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
    try:
        # Streaming obligatorio: con max_tokens grandes (cientos de filas) el
        # SDK rechaza la llamada no-streaming por el límite de 10 minutos.
        with client.messages.stream(
            model=settings.SAT_AI_MODEL,
            max_tokens=32000,
            system=[{"type": "text", "text": _SYSTEM_EXTRACT,
                     "cache_control": {"type": "ephemeral"}}],
            tools=[_TOOL_EXTRACT],
            tool_choice={"type": "tool", "name": "registrar_productos"},
            messages=[{"role": "user", "content": content}],
        ) as stream:
            resp = stream.get_final_message()
    except anthropic.APIError as exc:
        logger.warning("extracción IA de productos falló: %s", exc)
        raise ImportProductosError(
            "La IA no pudo leer el archivo en este momento. Intenta de nuevo o usa la plantilla."
        ) from exc

    filas: list[dict] = []
    for block in resp.content:
        if getattr(block, "type", None) == "tool_use" and block.name == "registrar_productos":
            for p in (block.input.get("productos") or []):
                if not isinstance(p, dict):
                    continue
                nombre = _texto(p.get("nombre"))
                if not nombre:
                    continue
                precio = _decimal(p.get("precio"))
                filas.append({
                    "nombre": nombre,
                    "codigo": _texto(p.get("codigo")),
                    "descripcion": _texto(p.get("descripcion")),
                    "unidad": normalizar_unidad(_texto(p.get("unidad"))),
                    "precio": str(precio) if precio is not None else "",
                    "clave_sat": re.sub(r"\D", "", _texto(p.get("clave_sat")))[:8],
                    "unidad_sat": _texto(p.get("unidad_sat")).upper()[:3],
                    "codigo_barras": "",
                    "categoria": _texto(p.get("categoria")),
                    "esquema": "",
                    "estatus": "",
                })
    if not filas:
        raise ImportProductosError("La IA no encontró productos en el archivo")
    return filas[:MAX_FILAS_IA]


def generar_plantilla() -> bytes:
    """La plantilla oficial .xlsx: encabezados + 2 filas de ejemplo + notas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F2937")
    for i, (nombre, _) in enumerate(PLANTILLA_COLUMNAS, start=1):
        c = ws.cell(row=1, column=i, value=nombre)
        c.font = head_font
        c.fill = head_fill
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(nombre) + 4)
    ws.column_dimensions["A"].width = 34

    ejemplos = [
        ["JITOMATE SALADETT", "JIT-SAD-001", "KILO", "28.50", "50406500", "KGM",
         "FRUTA Y VERDURA", "IVA0", "", "ALTA"],
        ["ACEITE COMESTIBLE 20 LT CRISTAL", "", "PIEZA", "935.40", "50151513", "H87",
         "ABARROTE", "IVA0", "", "ALTA"],
    ]
    for r, fila in enumerate(ejemplos, start=2):
        for i, v in enumerate(fila, start=1):
            ws.cell(row=r, column=i, value=v)

    notas = wb.create_sheet("Instrucciones")
    notas.column_dimensions["A"].width = 18
    notas.column_dimensions["B"].width = 70
    notas.cell(row=1, column=1, value="Columna").font = Font(bold=True)
    notas.cell(row=1, column=2, value="Qué va ahí").font = Font(bold=True)
    for r, (nombre, ayuda) in enumerate(PLANTILLA_COLUMNAS, start=2):
        notas.cell(row=r, column=1, value=nombre)
        notas.cell(row=r, column=2, value=ayuda)
    fila_extra = len(PLANTILLA_COLUMNAS) + 3
    notas.cell(row=fila_extra, column=2,
               value="Solo NOMBRE es obligatorio. Las filas de ejemplo se pueden borrar. "
                     "El SKU interno se genera automáticamente; lo que pongas en CODIGO "
                     "se guarda como código del cliente cuando importas la lista de un cliente.")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
