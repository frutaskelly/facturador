"""Remisiones end-to-end (Phase 4e): draft creation + folios, confirm reserving
stock, cancel releasing it, the almacén requirement, lifecycle guards, RBAC,
and isolation."""
import uuid

import pytest
from sqlalchemy import text

from app.core.auth import Principal, get_principal
from app.core.db import SessionLocal
from app.main import app
from app.models import Almacen, Cliente, Membership, Producto, Role, Tenant, User

_PURGE = (
    "movimientos_inventario", "mermas", "lineas_remision", "remisiones",
    "lotes_inventario", "precios", "listas_precios", "productos", "almacenes", "clientes",
)


@pytest.fixture
def env(db_engine):
    suffix = uuid.uuid4().hex[:8]
    db = SessionLocal()
    created = {"memberships": [], "users": [], "tenants": []}
    try:
        def _tenant(s):
            t = Tenant(slug=f"rem-{s}-{suffix}", legal_name=f"Rem {s} SA",
                       rfc=f"R{s.upper()}{suffix.upper()}"[:13], regimen_fiscal_sat="601",
                       domicilio_fiscal_cp="44100", tier="PRINCIPAL", status="ACTIVE")
            db.add(t); db.flush(); created["tenants"].append(t.id); return t

        tenant_a, tenant_b = _tenant("a"), _tenant("b")
        admin_role = db.query(Role).filter(Role.nombre == "ADMIN", Role.es_preset.is_(True)).one()
        tomador_role = db.query(Role).filter(Role.nombre == "TOMADOR", Role.es_preset.is_(True)).one()

        def _user(tenant, role, label):
            sub = f"sub-{label}-{suffix}"
            u = User(email=f"{label}-{suffix}@t.test", auth_user_id=sub, full_name=label)
            db.add(u); db.flush(); created["users"].append(u.id)
            m = Membership(tenant_id=tenant.id, user_id=u.id, role_id=role.id)
            db.add(m); db.flush(); created["memberships"].append(m.id)
            return {"sub": sub, "email": u.email, "tenant_id": tenant.id}

        admin_a = _user(tenant_a, admin_role, "admin-a")
        tomador_a = _user(tenant_a, tomador_role, "tomador-a")
        admin_b = _user(tenant_b, admin_role, "admin-b")

        cli = Cliente(tenant_id=tenant_a.id, codigo="CL1", legal_name="Cliente 1", rfc="XAXX010101000")
        prod = Producto(tenant_id=tenant_a.id, sku="R-P", nombre="Prod R", clave_sat="01010101", unidad_sat="KGM")
        prod_bulto = Producto(
            tenant_id=tenant_a.id, sku="R-PB", nombre="Prod Bulto R",
            clave_sat="50300000", unidad_sat="KGM",
            unidad_base="KILO", presentaciones={"KILO": 1, "BULTO": 20},
        )
        alm = Almacen(tenant_id=tenant_a.id, codigo="R-BG", nombre="Bodega R")
        db.add_all([cli, prod, alm, prod_bulto]); db.flush()
        db.commit()
        yield {"admin_a": admin_a, "tomador_a": tomador_a, "admin_b": admin_b,
               "cli_a": str(cli.id), "prod_a": str(prod.id), "alm_a": str(alm.id),
               "prod_bulto_a": str(prod_bulto.id)}
    finally:
        for table in _PURGE:
            for tid in created["tenants"]:
                db.execute(text(f"DELETE FROM {table} WHERE tenant_id = :tid"), {"tid": tid})
        for mid in created["memberships"]:
            db.query(Membership).filter(Membership.id == mid).delete()
        for uid in created["users"]:
            db.query(User).filter(User.id == uid).delete()
        for tid in created["tenants"]:
            db.query(Tenant).filter(Tenant.id == tid).delete()
        db.commit(); db.close()


@pytest.fixture
def auth_as():
    def _set(user):
        app.dependency_overrides[get_principal] = lambda: Principal(
            auth_user_id=user["sub"], email=user["email"], role="authenticated",
            claims={"sub": user["sub"]})
    yield _set
    app.dependency_overrides.pop(get_principal, None)


def _hdr(u):
    return {"X-Tenant-Id": str(u["tenant_id"])}


def _load_stock(client, h, env, qty, costo):
    return client.post("/api/v1/inventario/movimientos", headers=h, json={
        "tipo": "ENTRADA_COMPRA", "producto_id": env["prod_a"], "almacen_id": env["alm_a"],
        "cantidad": qty, "costo_unitario": costo})


def _create_rem(client, h, env, qty, precio, *, almacen=True):
    body = {"cliente_facturacion_id": env["cli_a"],
            "lineas": [{"producto_id": env["prod_a"], "cantidad_solicitada": qty, "precio_unitario": precio}]}
    if almacen:
        body["almacen_id"] = env["alm_a"]
    return client.post("/api/v1/remisiones", headers=h, json=body)


def _disp(client, h, env):
    rows = client.get("/api/v1/inventario/existencias", headers=h, params={"producto_id": env["prod_a"]}).json()
    return next((r for r in rows if r["almacen_id"] == env["alm_a"]), None)


def test_create_draft_and_folio(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    r = _create_rem(client, h, env, "10", "5")
    assert r.status_code == 201, r.text
    rem = r.json()
    assert rem["folio_interno"] == "R1"
    assert rem["estado"] == "BORRADOR"
    assert float(rem["subtotal"]) == 50.0
    assert float(rem["total"]) == 50.0
    assert len(rem["lineas"]) == 1 and rem["lineas"][0]["numero_linea"] == 1
    assert _create_rem(client, h, env, "1", "1").json()["folio_interno"] == "R2"


def test_confirm_descuenta_then_cancel_restituye(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    _load_stock(client, h, env, "100", "4")
    rem_id = _create_rem(client, h, env, "30", "5").json()["id"]

    c = client.post(f"/api/v1/remisiones/{rem_id}/confirmar", headers=h)
    assert c.status_code == 200, c.text
    assert c.json()["estado"] == "CONFIRMADA"
    row = _disp(client, h, env)
    assert float(row["disponible"]) == 70.0
    assert float(row["reservada"]) == 0.0   # salida directa: sin cubeta de apartado
    movs = client.get("/api/v1/inventario/movimientos", headers=h, params={"tipo": "SALIDA_REMISION"}).json()
    assert movs["total"] >= 1

    x = client.post(f"/api/v1/remisiones/{rem_id}/cancelar", headers=h)
    assert x.status_code == 200 and x.json()["estado"] == "CANCELADA"
    row2 = _disp(client, h, env)
    assert float(row2["disponible"]) == 100.0
    assert float(row2["reservada"]) == 0.0


def test_confirm_with_presentation_descuenta_base_units(client, env, auth_as):
    """Selling in BULTO (1 BULTO = 20 KILO) descuenta el equivalente en unidad
    base: 100 KILO en stock, 2 BULTO → salen 40 KILO. Cancelar restituye 40."""
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    pid = env["prod_bulto_a"]
    client.post("/api/v1/inventario/movimientos", headers=h, json={
        "tipo": "ENTRADA_COMPRA", "producto_id": pid, "almacen_id": env["alm_a"],
        "cantidad": "100", "costo_unitario": "4"})
    rem_id = client.post("/api/v1/remisiones", headers=h, json={
        "cliente_facturacion_id": env["cli_a"], "almacen_id": env["alm_a"],
        "lineas": [{"producto_id": pid, "cantidad_solicitada": "2",
                    "precio_unitario": "150", "presentacion": "BULTO"}]}).json()["id"]

    def _row():
        rows = client.get("/api/v1/inventario/existencias", headers=h, params={"producto_id": pid}).json()
        return next(r for r in rows if r["almacen_id"] == env["alm_a"])

    assert client.post(f"/api/v1/remisiones/{rem_id}/confirmar", headers=h).status_code == 200
    row = _row()
    assert float(row["disponible"]) == 60.0   # 100 − (2 × 20)
    assert float(row["reservada"]) == 0.0   # salida directa: sin cubeta de apartado

    assert client.post(f"/api/v1/remisiones/{rem_id}/cancelar", headers=h).status_code == 200
    row2 = _row()
    assert float(row2["disponible"]) == 100.0
    assert float(row2["reservada"]) == 0.0


def test_auto_precio_desde_lista(client, env, auth_as):
    """Sin precio en la línea → se resuelve desde la lista base (UNICO)."""
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    lista = client.post("/api/v1/listas-precios", headers=h, json={"codigo": "UNICO", "nombre": "Único"}).json()
    client.post(f"/api/v1/listas-precios/{lista['id']}/precios", headers=h,
                json={"producto_id": env["prod_a"], "precio_unitario": "7.50", "cantidad_minima": 1})
    r = client.post("/api/v1/remisiones", headers=h, json={
        "cliente_facturacion_id": env["cli_a"], "almacen_id": env["alm_a"],
        "lineas": [{"producto_id": env["prod_a"], "cantidad_solicitada": "4"}]})  # sin precio_unitario
    assert r.status_code == 201, r.text
    rem = r.json()
    assert float(rem["lineas"][0]["precio_unitario"]) == 7.50
    assert float(rem["subtotal"]) == 30.0  # 4 × 7.50


def test_auto_precio_sin_precio_disponible_422(client, env, auth_as):
    """Sin precio en línea ni lista ni override → 422 (pide precio manual)."""
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    r = client.post("/api/v1/remisiones", headers=h, json={
        "cliente_facturacion_id": env["cli_a"], "almacen_id": env["alm_a"],
        "lineas": [{"producto_id": env["prod_a"], "cantidad_solicitada": "1"}]})
    assert r.status_code == 422


def test_confirm_with_real_weight_catch_weight(client, env, auth_as):
    """Confirmar con peso real por línea (catch-weight): descuenta 43 kg (no el
    estimado 40 = 2×20), y cancelar restituye exactamente esos 43."""
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    pid = env["prod_bulto_a"]
    client.post("/api/v1/inventario/movimientos", headers=h, json={
        "tipo": "ENTRADA_COMPRA", "producto_id": pid, "almacen_id": env["alm_a"],
        "cantidad": "100", "costo_unitario": "4"})
    rem = client.post("/api/v1/remisiones", headers=h, json={
        "cliente_facturacion_id": env["cli_a"], "almacen_id": env["alm_a"],
        "lineas": [{"producto_id": pid, "cantidad_solicitada": "2",
                    "precio_unitario": "150", "presentacion": "BULTO"}]}).json()
    rem_id, linea_id = rem["id"], rem["lineas"][0]["id"]

    def _row():
        rows = client.get("/api/v1/inventario/existencias", headers=h, params={"producto_id": pid}).json()
        return next(r for r in rows if r["almacen_id"] == env["alm_a"])

    c = client.post(f"/api/v1/remisiones/{rem_id}/confirmar", headers=h,
                    json={"pesos": [{"linea_id": linea_id, "cantidad_base": "43"}]})
    assert c.status_code == 200, c.text
    assert float(_row()["disponible"]) == 57.0   # 100 − 43 (real, no 40)
    assert float(_row()["reservada"]) == 0.0   # salida directa: sin cubeta de apartado

    assert client.post(f"/api/v1/remisiones/{rem_id}/cancelar", headers=h).status_code == 200
    assert float(_row()["disponible"]) == 100.0 and float(_row()["reservada"]) == 0.0


def test_confirm_insufficient_stock(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    rem_id = _create_rem(client, h, env, "30", "5").json()["id"]  # no stock loaded
    assert client.post(f"/api/v1/remisiones/{rem_id}/confirmar", headers=h).status_code == 422


def test_confirm_requires_almacen(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    rem_id = _create_rem(client, h, env, "5", "5", almacen=False).json()["id"]
    assert client.post(f"/api/v1/remisiones/{rem_id}/confirmar", headers=h).status_code == 422


def test_cancel_draft_and_guard(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    rem_id = _create_rem(client, h, env, "5", "5").json()["id"]
    assert client.post(f"/api/v1/remisiones/{rem_id}/cancelar", headers=h).json()["estado"] == "CANCELADA"
    # can't confirm a cancelled remisión
    assert client.post(f"/api/v1/remisiones/{rem_id}/confirmar", headers=h).status_code == 409


def test_tomador_cannot_touch_remisiones(client, env, auth_as):
    auth_as(env["tomador_a"]); h = _hdr(env["tomador_a"])  # no menu:remisiones
    assert client.get("/api/v1/remisiones", headers=h).status_code == 403
    assert _create_rem(client, h, env, "1", "1").status_code == 403


def test_remisiones_isolated_between_tenants(client, env, auth_as):
    auth_as(env["admin_a"]); ha = _hdr(env["admin_a"])
    rem_id = _create_rem(client, ha, env, "1", "1").json()["id"]

    auth_as(env["admin_b"]); hb = _hdr(env["admin_b"])
    assert client.get(f"/api/v1/remisiones/{rem_id}", headers=hb).status_code == 404
    # B referencing tenant A's cliente → ensure_fk 422
    cross = client.post("/api/v1/remisiones", headers=hb, json={
        "cliente_facturacion_id": env["cli_a"],
        "lineas": [{"producto_id": env["prod_a"], "cantidad_solicitada": "1", "precio_unitario": "1"}]})
    assert cross.status_code == 422


# ── Importación masiva estilo SAE (Excel → varias remisiones) ────────────────
def _xlsx_sae(rows):
    """Construye un .xlsx en memoria con el layout SAE (FOLIO/CLIENTE/…)."""
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["FOLIO", "CLIENTE", "FECHA", "SU PEDIDO", "CLAVE", "CANTIDAD", "PRECIO", "Observaciones"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_importar_preview_agrupa_y_cruza(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    sku = client.get("/api/v1/productos", headers=h, params={"limit": 1}).json()["items"][0]
    data = _xlsx_sae([
        ["0000001230", "C-X", "29/07/2026", "OC 1", sku["sku"], "2", "10", "entrega lunes"],
        ["0000001230", "C-X", "29/07/2026", None, "SKU-QUE-NO-EXISTE", "1", "5", None],
        ["ZHGO9",      "C-X", "29/07/2026", None, sku["sku"], "3", "12", None],
    ])
    r = client.post("/api/v1/remisiones/importar-preview", headers=h,
                    files={"archivo": ("pedidos.xlsx", data,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    p = r.json()
    assert len(p["grupos"]) == 2
    g1 = next(g for g in p["grupos"] if g["folio_ref"] == "1230")   # sin ceros
    assert g1["su_pedido"] == "OC 1"
    assert len(g1["lineas"]) == 2
    cruzada = next(l for l in g1["lineas"] if l["clave"] == sku["sku"])
    assert cruzada["producto_id"] == sku["id"]                       # CLAVE = SKU exacto
    sin_cruce = next(l for l in g1["lineas"] if l["clave"] == "SKU-QUE-NO-EXISTE")
    assert sin_cruce["producto_id"] is None
    assert p["productos_sin_cruce"] == 1


def test_importar_preview_rechaza_formato_invalido(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    r = client.post("/api/v1/remisiones/importar-preview", headers=h,
                    files={"archivo": ("nota.txt", b"esto no es excel", "text/plain")})
    assert r.status_code == 422


# ── Importación del Master Ordenes (hoja "Master" del concentrado de OC) ─────
def _xlsx_master(rows):
    """Construye un .xlsx con el layout del Master Ordenes: la hoja de renglones
    se llama "Master" y el libro arrastra además "Summary"/"Totales"."""
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Master"
    ws.append([
        "Archivo", "Tipo Documento", "RFC Cliente", "Nombre Cliente", "RFC Proveedor",
        "Nombre Proveedor", "Folio", "Requisicion Folio", "Referencia", "Fecha",
        "Cantidad", "Unidad", "Clave", "Descripcion", "Costo unitario", "DESC",
        "Subtotal", "Observacion del documento", "Entregar Bodega",
    ])
    for r in rows:
        ws.append(r)
    wb.create_sheet("Summary").append(["Folio", "Cliente", "Total"])
    wb.create_sheet("Totales").append(["Documentos", len(rows)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_importar_master_ordenes_cruza_por_rfc_y_nombre(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    sku = client.get("/api/v1/productos", headers=h, params={"limit": 1}).json()["items"][0]
    fila = lambda folio, rfc, nombre, clave, cant: [  # noqa: E731
        "OCO 458.pdf", "ORDEN DE COMPRA", rfc, nombre, "ZAOC830517RF9",
        "CRISTIAN GERARDO ZARATE OROZCO", folio, "0000000271", "CEUHM VERDURA",
        "46209",  # serial de Excel = 2026-07-06
        cant, "KILOGRAMO", clave, "AJO MORADO", "90", "0", "45",
        "ENTREGAR EN COMEDOR EL JUEVES", "16 DE JUL",
    ]
    data = _xlsx_master([
        fila("0000000458", "XAXX010101000", "Cliente 1", sku["sku"], "2"),
        fila("0000000458", "XAXX010101000", "Cliente 1", "AJO -FRUT-017", "0.5"),
        # RFC desconocido: el cruce cae al nombre, sin la razón social.
        fila("0000024460", "AAA010101AAA", "Cliente 1 S.A. de C.V.", sku["sku"], "3"),
    ])
    r = client.post("/api/v1/remisiones/importar-preview", headers=h,
                    files={"archivo": ("Master Ordenes.xlsx", data,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    p = r.json()
    assert len(p["grupos"]) == 2
    g1 = next(g for g in p["grupos"] if g["folio_ref"] == "458")     # sin ceros
    assert g1["cliente_id"] == env["cli_a"]                          # cruzó por RFC
    assert g1["fecha"] == "2026-07-06"                               # serial de Excel
    assert g1["su_pedido"] == "CEUHM VERDURA"                        # Referencia
    assert g1["observaciones"] == "ENTREGAR EN COMEDOR EL JUEVES"
    assert g1["requisicion"] == "271"                                # sin ceros
    assert g1["entregar_bodega"] == "16 DE JUL"
    cruzada = next(l for l in g1["lineas"] if l["clave"] == sku["sku"])
    assert cruzada["producto_id"] == sku["id"] and float(cruzada["precio"]) == 90
    sin_cruce = next(l for l in g1["lineas"] if l["clave"] == "AJO -FRUT-017")
    assert sin_cruce["producto_id"] is None
    assert sin_cruce["descripcion"] == "AJO MORADO" and sin_cruce["unidad"] == "KILOGRAMO"
    g2 = next(g for g in p["grupos"] if g["folio_ref"] == "24460")
    assert g2["cliente_id"] == env["cli_a"]                          # cruzó por nombre
    assert p["clientes_sin_cruce"] == 0


# ── Factura de SAE → estado RESERVADO ─────────────────────────────────────────
def test_factura_sae_reserva_y_al_quitarla_vuelve_a_borrador(client, env, auth_as):
    auth_as(env["admin_a"]); h = _hdr(env["admin_a"])
    base = {"cliente_facturacion_id": env["cli_a"], "almacen_id": env["alm_a"],
            "lineas": [{"producto_id": env["prod_a"], "cantidad_solicitada": "2", "precio_unitario": "10"}]}

    # Nace RESERVADA si el alta trae el folio de SAE.
    r = client.post("/api/v1/remisiones", headers=h,
                    json={**base, "factura_sae": "ZHGO 233", "su_pedido": "0000024478"})
    assert r.status_code == 201, r.text
    assert r.json()["estado"] == "RESERVADO"
    assert r.json()["factura_sae"] == "ZHGO 233"
    # "Su pedido" es la OC del cliente y va en su propia columna, no en las notas.
    assert r.json()["su_pedido"] == "0000024478"

    # Sin folio nace BORRADOR; ponérselo después la reserva.
    rem = client.post("/api/v1/remisiones", headers=h, json=base).json()
    assert rem["estado"] == "BORRADOR" and rem["factura_sae"] is None
    up = client.patch(f"/api/v1/remisiones/{rem['id']}", headers=h, json={"factura_sae": "ZHGO 234"})
    assert up.status_code == 200, up.text
    assert up.json()["estado"] == "RESERVADO"

    # Quitarlo la regresa a BORRADOR.
    up = client.patch(f"/api/v1/remisiones/{rem['id']}", headers=h, json={"factura_sae": ""})
    assert up.json()["estado"] == "BORRADOR" and up.json()["factura_sae"] is None

    # Una RESERVADA se confirma igual que un borrador (la salida es el confirmar).
    client.post("/api/v1/inventario/movimientos", headers=h, json={
        "tipo": "ENTRADA_COMPRA", "producto_id": env["prod_a"], "almacen_id": env["alm_a"],
        "cantidad": "10", "costo_unitario": "4"})
    client.patch(f"/api/v1/remisiones/{rem['id']}", headers=h, json={"factura_sae": "ZHGO 235"})
    conf = client.post(f"/api/v1/remisiones/{rem['id']}/confirmar", headers=h, json={})
    assert conf.status_code == 200, conf.text
    assert conf.json()["estado"] == "CONFIRMADA"
    # Ya confirmada, el folio de SAE es un dato más: no la regresa a borrador.
    up = client.patch(f"/api/v1/remisiones/{rem['id']}", headers=h, json={"factura_sae": ""})
    assert up.json()["estado"] == "CONFIRMADA"
